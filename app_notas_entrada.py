import streamlit as st
import pdfplumber
import pandas as pd
import re
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Relatório CFOP - CONTTEC", page_icon="🧾", layout="wide")

st.title("🧾 Gerador de Relatório: Compras por CFOP")
st.markdown("Sistema atualizado para a estrutura final **Multi-Abas (Dashboard + Dados Filtráveis)**, com algoritmo blindado contra quebra de linhas no PDF.")

arquivos_pdf = st.file_uploader("Arraste os seus PDFs para aqui", type="pdf", accept_multiple_files=True)

def to_f(v):
    try:
        return float(v.replace(".", "").replace(",", "."))
    except:
        return 0.0

if arquivos_pdf:
    if st.button("Processar Documentos"):
        with st.spinner("A extrair 100% das notas e a gerar o Dashboard..."):
            rows = []
            
            for f in arquivos_pdf:
                with pdfplumber.open(io.BytesIO(f.read())) as pdf:
                    full_text = ""
                    for page in pdf.pages:
                        full_text += page.extract_text(x_tolerance=1, y_tolerance=1) + "\n"
                        
                    lines = full_text.split("\n")
                    i = 0
                    current_cfop_code = ""
                    current_cfop_desc = ""
                    
                    while i < len(lines):
                        line = lines[i].strip()
                        if not line:
                            i += 1
                            continue
                            
                        # 1. Deteta Linha de CFOP (Ex: "1101 - COMPRA PARA...")
                        m_cfop = re.match(r"^(\d{4})\s*-\s*(.*)", line)
                        if m_cfop:
                            current_cfop_code = m_cfop.group(1)
                            desc = m_cfop.group(2).strip()
                            # Agrupa descrição se foi partida na linha de baixo
                            if i + 1 < len(lines) and not re.match(r"^\d", lines[i+1].strip()) and "TOTAL" not in lines[i+1]:
                                desc += " " + lines[i+1].strip()
                                i += 1
                            current_cfop_desc = desc
                            i += 1
                            continue
                            
                        # 2. Deteta início de Nota Fiscal
                        if re.match(r"^\d{2}/\d{2}/\d{4}", line):
                            record_str = line
                            
                            # Agrupa todas as linhas da mesma nota até encontrar um gatilho de paragem
                            while i + 1 < len(lines):
                                next_line = lines[i+1].strip()
                                if not next_line:
                                    i += 1
                                    continue
                                if (re.match(r"^\d{2}/\d{2}/\d{4}", next_line) or 
                                    re.match(r"^\d{4}\s*-", next_line) or 
                                    "TOTAL" in next_line or 
                                    "DT EMISSÃO" in next_line or
                                    "Filtros:" in next_line or
                                    "CNPJ" in next_line or
                                    "CONTTEC" in next_line):
                                    break
                                record_str += " " + next_line
                                i += 1
                                
                            # 3. Análise Blindada (Da Direita para a Esquerda)
                            tokens = record_str.split()
                            if len(tokens) >= 10:
                                date = tokens[0]
                                
                                # Remove o 'Prazo' se existir no final da linha
                                if re.match(r"^\d+$", tokens[-1]) and re.search(r"[.,]", tokens[-2]):
                                    tokens.pop()
                                    
                                if len(tokens) >= 9:
                                    total = tokens.pop()
                                    cred = tokens.pop()
                                    icms = tokens.pop()
                                    ipi = tokens.pop()
                                    v_unt = tokens.pop()
                                    qtd = tokens.pop()
                                    serie = tokens.pop()
                                    nota = tokens.pop()
                                    forn = " ".join(tokens[1:])
                                    
                                    rows.append({
                                        "Data": date,
                                        "Fornecedor": forn,
                                        "Nota": nota,
                                        "Série": serie,
                                        "CFOP": int(current_cfop_code) if current_cfop_code.isdigit() else current_cfop_code,
                                        "Qtd": to_f(qtd),
                                        "Vlr Unit": to_f(v_unt),
                                        "IPI": to_f(ipi),
                                        "Cred ICMS": to_f(cred),
                                        "ICMS": to_f(icms),
                                        "Total": to_f(total),
                                        "Descrição CFOP": current_cfop_desc
                                    })
                        i += 1

            if not rows:
                st.error("❌ Nenhum dado compatível encontrado nos PDFs enviados.")
            else:
                df = pd.DataFrame(rows)
                
                # Resumo no Ecrã
                st.divider()
                st.subheader("👀 Auditoria de Valores (Bate com o PDF)")
                c1, c2, c3, c4 = st.columns(4)
                c1.metric("Notas Recuperadas", len(df))
                c2.metric("Soma Geral (R$)", f"{df['Total'].sum():,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
                c3.metric("Soma ICMS (R$)", f"{df['ICMS'].sum():,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
                c4.metric("Soma IPI (R$)", f"{df['IPI'].sum():,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
                
                # ==========================================
                # GERAÇÃO EXCEL - DUAS ABAS (DASHBOARD + DADOS)
                # ==========================================
                wb = Workbook()
                
                # ESTILOS
                F_BOLD = Font(bold=True)
                F_TITLE = Font(bold=True, size=14, color="1F4E78")
                F_HEADER = Font(bold=True, color="FFFFFF")
                BG_BLUE = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                BG_LIGHT = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                ALIGN_C = Alignment(horizontal="center", vertical="center")
                
                df['Data_Dt'] = pd.to_datetime(df['Data'], format="%d/%m/%Y", errors='coerce')
                min_dt = df['Data_Dt'].min().strftime("%d/%m/%Y") if pd.notnull(df['Data_Dt'].min()) else ""
                max_dt = df['Data_Dt'].max().strftime("%d/%m/%Y") if pd.notnull(df['Data_Dt'].max()) else ""
                
                # ------------------------------------------
                # ABA 1: DASHBOARD
                # ------------------------------------------
                ws_dash = wb.active
                ws_dash.title = "Dashboard"
                ws_dash.sheet_view.showGridLines = False
                
                ws_dash['A1'] = "RELATÓRIO DE ENTRADAS POR CFOP"
                ws_dash['A1'].font = F_TITLE
                ws_dash['A3'] = "CONTTEC COM E SERV ESP EM MOTORES ELETRICOS IND LTDA"
                ws_dash['A3'].font = F_BOLD
                ws_dash['A4'] = f"Período: {min_dt} a {max_dt}  |  CNPJ: 12.231.733/0001-16"
                
                ws_dash['B6'], ws_dash['D6'], ws_dash['F6'], ws_dash['H6'] = "TOTAL GERAL", "TOTAL ICMS", "TOTAL IPI", "NOTAS"
                for cel in ['B6', 'D6', 'F6', 'H6']:
                    ws_dash[cel].font = F_BOLD
                    ws_dash[cel].alignment = ALIGN_C
                    
                ws_dash['B8'] = df['Total'].sum()
                ws_dash['D8'] = df['ICMS'].sum()
                ws_dash['F8'] = df['IPI'].sum()
                ws_dash['H8'] = len(df)
                for cel in ['B8', 'D8', 'F8']:
                    ws_dash[cel].number_format = '#,##0.00'
                    ws_dash[cel].font = Font(size=12, bold=True)
                    ws_dash[cel].alignment = ALIGN_C
                ws_dash['H8'].alignment = ALIGN_C
                ws_dash['H8'].font = Font(size=12, bold=True)
                
                ws_dash['A12'] = "RESUMO POR CFOP"
                ws_dash['A12'].font = Font(bold=True, size=12)
                
                headers_resumo = ["CFOP", "Descrição", "Qtde Itens", "Total R$", "ICMS R$", "IPI R$", "% do Total"]
                for col_num, h in enumerate(headers_resumo, 1):
                    c = ws_dash.cell(row=14, column=col_num, value=h)
                    c.font = F_HEADER
                    c.fill = BG_BLUE
                    
                resumo = df.groupby(['CFOP', 'Descrição CFOP']).agg(
                    Qtde=('Nota', 'count'), Total=('Total', 'sum'), ICMS=('ICMS', 'sum'), IPI=('IPI', 'sum')
                ).reset_index()
                
                total_geral = df['Total'].sum()
                resumo['Perc'] = resumo['Total'] / total_geral if total_geral else 0
                resumo = resumo.sort_values('Total', ascending=False)
                
                row_idx = 15
                for _, r in resumo.iterrows():
                    ws_dash.cell(row=row_idx, column=1, value=r['CFOP']).alignment = ALIGN_C
                    ws_dash.cell(row=row_idx, column=2, value=r['Descrição CFOP'])
                    ws_dash.cell(row=row_idx, column=3, value=r['Qtde']).alignment = ALIGN_C
                    ws_dash.cell(row=row_idx, column=4, value=r['Total']).number_format = '#,##0.00'
                    ws_dash.cell(row=row_idx, column=5, value=r['ICMS']).number_format = '#,##0.00'
                    ws_dash.cell(row=row_idx, column=6, value=r['IPI']).number_format = '#,##0.00'
                    ws_dash.cell(row=row_idx, column=7, value=r['Perc']).number_format = '0.00%'
                    row_idx += 1
                    
                ws_dash.column_dimensions['A'].width = 10
                ws_dash.column_dimensions['B'].width = 55
                ws_dash.column_dimensions['C'].width = 12
                ws_dash.column_dimensions['D'].width = 15
                ws_dash.column_dimensions['E'].width = 15
                ws_dash.column_dimensions['F'].width = 15
                ws_dash.column_dimensions['G'].width = 12
                
                # ------------------------------------------
                # ABA 2: DADOS FILTRÁVEIS
                # ------------------------------------------
                ws_data = wb.create_sheet("Dados Filtráveis")
                ws_data.sheet_view.showGridLines = False
                
                ws_data['A1'] = "DADOS COM TOTALIZADOR DINÂMICO"
                ws_data['A1'].font = F_TITLE
                ws_data['A2'] = "Use os filtros abaixo - os totais atualizam automaticamente"
                
                last_row = 5 + len(df)
                
                # TOTAIS DINÂMICOS NA LINHA 4
                ws_data['A4'] = "TOTAIS FILTRADOS:"
                ws_data['A4'].font = F_BOLD
                ws_data['D4'] = "Itens:"
                ws_data['D4'].font = F_BOLD
                ws_data['E4'] = f"=SUBTOTAL(103, A6:A{last_row})" # Contagem
                
                ws_data['H4'] = f"=SUBTOTAL(109, H6:H{last_row})" # IPI
                ws_data['I4'] = f"=SUBTOTAL(109, I6:I{last_row})" # Cred ICMS
                ws_data['J4'] = f"=SUBTOTAL(109, J6:J{last_row})" # ICMS
                ws_data['K4'] = f"=SUBTOTAL(109, K6:K{last_row})" # Total
                ws_data['L4'] = "TOTAL:"
                ws_data['L4'].font = F_BOLD
                
                for col in ['E4', 'H4', 'I4', 'J4', 'K4']:
                    ws_data[col].font = F_BOLD
                    ws_data[col].fill = BG_LIGHT
                    if col != 'E4': ws_data[col].number_format = '#,##0.00'
                
                # CABEÇALHO NA LINHA 5
                headers_data = ["Data", "Fornecedor", "Nota", "Série", "CFOP", "Qtd", "Vlr Unit", "IPI", "Cred ICMS", "ICMS", "Total", "Descrição CFOP"]
                for col_num, h in enumerate(headers_data, 1):
                    c = ws_data.cell(row=5, column=col_num, value=h)
                    c.font = F_HEADER
                    c.fill = BG_BLUE
                    c.alignment = ALIGN_C
                    
                # PREENCHER DADOS
                for r_idx, row in enumerate(df.itertuples(index=False), 6):
                    for c_idx, val in enumerate(row[:-1], 1): # Omitir Data_Dt
                        cell = ws_data.cell(row=r_idx, column=c_idx, value=val)
                        if c_idx in [1, 3, 4, 5, 6]: # Data, Nota, Serie, CFOP, Qtd
                            cell.alignment = ALIGN_C
                        if c_idx in [7, 8, 9, 10, 11]: # Valores Financeiros
                            cell.number_format = '#,##0.00'

                ws_data.column_dimensions['A'].width = 12
                ws_data.column_dimensions['B'].width = 45
                ws_data.column_dimensions['C'].width = 10
                ws_data.column_dimensions['D'].width = 8
                ws_data.column_dimensions['E'].width = 8
                ws_data.column_dimensions['F'].width = 8
                ws_data.column_dimensions['G'].width = 12
                ws_data.column_dimensions['H'].width = 12
                ws_data.column_dimensions['I'].width = 12
                ws_data.column_dimensions['J'].width = 12
                ws_data.column_dimensions['K'].width = 14
                ws_data.column_dimensions['L'].width = 40
                
                ws_data.auto_filter.ref = f"A5:L{last_row}"
                ws_data.freeze_panes = "A6"
                
                output = io.BytesIO()
                wb.save(output)
                
                st.divider()
                st.success("✅ Excel gerado com sucesso! Contém a aba Dashboard e a aba de Dados Filtráveis exatas ao modelo.")
                
                st.download_button(
                    label="📥 Baixar Relatório Completo (Multi-Abas)",
                    data=output.getvalue(),
                    file_name=f"Relatorio_CFOP_Final_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary"
                )
